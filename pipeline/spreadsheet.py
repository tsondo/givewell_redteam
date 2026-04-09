"""CEA spreadsheet readers and sensitivity analysis for GiveWell interventions.

Provides WaterCEA (water chlorination), ITNCEA (insecticide-treated nets),
and MalariaCEA (SMC) classes that replicate critical formula chains in Python
and run sensitivity analysis when parameters change.
"""
from __future__ import annotations

import math
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Protocol, runtime_checkable

import openpyxl


# ---------------------------------------------------------------------------
# CEA reader contract
# ---------------------------------------------------------------------------

@runtime_checkable
class CEAReader(Protocol):
    """Interface contract for CEA spreadsheet readers.

    Every concrete CEA class consumed by the pipeline must implement these
    members. The pipeline's quantifier stage iterates ``PROGRAMS`` to run
    cross-program sensitivity analysis; missing this attribute caused a
    mid-run crash on the VAS pipeline (quantifier critique 5/31, fixed in
    a prior commit). This Protocol exists to surface that class of bug at
    test time rather than at run time.

    Note: this is a structural Protocol (PEP 544), not an ABC. Existing
    classes conform automatically without inheriting — we only use it for
    isinstance checks in tests.
    """

    PROGRAMS: tuple[str, ...]

    def get_parameter_summary(self) -> str:
        """Return a markdown-formatted summary of CEA parameters for LLM agents."""
        ...

    def compute_cost_effectiveness(
        self, program_key: str, **overrides: float
    ) -> float:
        """Return the CE multiple for a program, optionally with parameter overrides.

        Implementations that read pre-computed values (rather than replicating
        the formula chain) should raise NotImplementedError if overrides are
        passed, to fail loud rather than silently returning stale values.
        """
        ...

    def detect_cap_binding(self, program_key: str, **overrides: float) -> bool:
        """Return True if the plausibility cap binds at the given parameter values."""
        ...

    def run_sensitivity(
        self,
        program_key: str,
        parameter_name: str,
        low: float,
        central: float,
        high: float,
    ) -> dict[str, Any]:
        """Run sensitivity analysis on a parameter and return structured results."""
        ...


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class ProgramParams:
    """Country/program-specific parameters read from the spreadsheet."""

    name: str
    # Demographics
    pop_under5: float
    pop_5_14: float
    baseline_mortality_under5: float
    baseline_mortality_over5: float

    # Validity
    external_validity: float
    plausibility_cap: float
    morbidity_ext_validity: float

    # Moral weights
    moral_weight_under5: float
    moral_weight_over5: float
    moral_weight_yld: float  # always 2.3 but read from sheet

    # Adult mortality
    adult_mortality_scaling: float

    # Morbidity
    ylds_per_100k: float
    yld_under5: float
    yld_5_14: float

    # Development effects
    dev_base: float  # units of value per treatment-year from SMC
    dev_chlorination_pct: float  # chlorination as % of SMC

    # Medical costs averted
    baseline_diarrhea_incidence: float
    cost_per_under5_diarrhea: float
    consumption: float  # annual consumption per capita
    iv_medical: float
    ev_medical: float
    value_doubling: float
    mills_reincke: float  # shared but read per-program for safety

    # Cost
    cost_per_person: float
    cash_benchmark: float

    # DSW adjustment factors (only used for DSW programs)
    downside_factor: float = 1.0
    excluded_effects_factor: float = 1.0

    # DSW leverage/funging
    has_leverage_adjustment: bool = False
    prob_govt_replace: float = 0.0
    prob_other_phil_replace: float = 0.0
    prob_upstream_stays: float = 0.0
    prob_not_exist: float = 0.0
    frac_govt_replace: float = 1.0
    frac_other_phil: float = 1.0
    frac_upstream: float = 0.0
    frac_not_exist: float = 0.0
    ce_govt: float = 0.0
    ce_other_phil: float = 0.0
    govt_costs: float = 0.0
    other_donor_upstream: float = 0.0


@dataclass
class SensitivityResult:
    """Result of a sensitivity analysis run."""

    parameter_name: str
    baseline: float
    low: float
    central: float
    high: float
    pct_change_low: float
    pct_change_central: float
    pct_change_high: float
    cap_binds_baseline: bool
    cap_binds_low: bool
    cap_binds_central: bool
    cap_binds_high: bool


# ---------------------------------------------------------------------------
# Main class
# ---------------------------------------------------------------------------

class WaterCEA:
    """Reads GiveWell Water Chlorination CEA and replicates the formula chain."""

    PROGRAMS = ("ilc_kenya", "dsw_kenya", "dsw_uganda", "dsw_malawi")

    def __init__(self, path: Path) -> None:
        self._path = Path(path)
        wb = openpyxl.load_workbook(str(self._path), data_only=True)
        self._read_shared_params(wb)
        self._read_program_params(wb)
        wb.close()

    # ------------------------------------------------------------------
    # Reading helpers
    # ------------------------------------------------------------------

    def _cell(self, ws: Any, ref: str) -> Any:
        """Read a cell value, raising if None."""
        val = ws[ref].value
        if val is None:
            raise ValueError(f"Cell {ws.title}!{ref} is None")
        return val

    def _read_shared_params(self, wb: openpyxl.Workbook) -> None:
        mort = wb["Mortality effect size"]
        self.pooled_ln_rr: float = float(self._cell(mort, "B10"))
        self.relative_risk: float = float(self._cell(mort, "B11"))

        iv = wb["Internal validity adjustment"]
        self.internal_validity_under5: float = float(self._cell(iv, "B6"))
        self.internal_validity_over5: float = float(self._cell(iv, "B7"))
        self.internal_validity_morbidity: float = float(self._cell(iv, "E3"))
        self.mills_reincke: float = float(self._cell(iv, "B23"))

        morb = wb["Morbidity effect size"]
        self.adjusted_diarrhea_rr: float = float(self._cell(morb, "B7"))

    def _read_program_params(self, wb: openpyxl.Workbook) -> None:
        self.programs: dict[str, ProgramParams] = {}
        self._read_ilc_kenya(wb)
        self._read_dsw(wb, "dsw_kenya", "B")
        self._read_dsw(wb, "dsw_uganda", "C")
        self._read_dsw(wb, "dsw_malawi", "D")

    def _read_ilc_kenya(self, wb: openpyxl.Workbook) -> None:
        ws = wb["ILC Kenya"]
        ev_ws = wb["External validity adjustment"]
        mw = wb["Moral weights"]
        am = wb["Adult mortality scaling factor"]
        iv = wb["Internal validity adjustment"]

        self.programs["ilc_kenya"] = ProgramParams(
            name="ILC Kenya",
            pop_under5=float(self._cell(ws, "B3")),
            pop_5_14=float(self._cell(ws, "B42")),
            baseline_mortality_under5=float(self._cell(ws, "B4")),
            baseline_mortality_over5=float(self._cell(ws, "B17")),
            external_validity=float(self._cell(ev_ws, "B6")),
            plausibility_cap=float(self._cell(iv, "B11")),
            morbidity_ext_validity=float(self._cell(ev_ws, "B29")),
            moral_weight_under5=float(self._cell(mw, "B3")),
            moral_weight_over5=float(self._cell(mw, "B7")),
            moral_weight_yld=float(self._cell(ws, "B35")),
            adult_mortality_scaling=float(self._cell(am, "B17")),
            ylds_per_100k=float(self._cell(ws, "B29")),
            yld_under5=float(self._cell(ws, "B43")),
            yld_5_14=float(self._cell(ws, "B44")),
            dev_base=float(self._cell(ws, "B39")),
            dev_chlorination_pct=float(self._cell(ws, "B40")),
            baseline_diarrhea_incidence=float(self._cell(ws, "B52")),
            cost_per_under5_diarrhea=float(self._cell(ws, "B55")),
            consumption=float(self._cell(ws, "B63")),
            iv_medical=float(self._cell(ws, "B60")),
            ev_medical=float(self._cell(ws, "B61")),
            value_doubling=float(self._cell(ws, "B65")),
            mills_reincke=float(self._cell(ws, "B58")),
            cost_per_person=float(self._cell(ws, "B72")),
            cash_benchmark=float(self._cell(ws, "B77")),
        )

    def _read_dsw(self, wb: openpyxl.Workbook, key: str, col: str) -> None:
        ws = wb["DSW"]
        ev_ws = wb["External validity adjustment"]
        mw = wb["Moral weights"]
        am = wb["Adult mortality scaling factor"]
        iv = wb["Internal validity adjustment"]

        # Map column to country index for moral weights and adult mortality
        col_map = {"B": "B", "C": "C", "D": "D"}
        am_col = col_map[col]

        # Plausibility caps: B12=Kenya, B13=Uganda, B14=Malawi
        cap_row = {"B": 12, "C": 13, "D": 14}[col]

        # Moral weights: under5 = B3(Kenya), B4(Uganda), B5(Malawi)
        mw_u5_row = {"B": 3, "C": 4, "D": 5}[col]
        # over5 = B7(Kenya), B8(Uganda), B9(Malawi)
        mw_o5_row = {"B": 7, "C": 8, "D": 9}[col]

        # External validity for mortality
        ev_row = {"B": 11, "C": 16, "D": 21}[col]
        # External validity for morbidity
        morb_ev_row = {"B": 33, "C": 37, "D": 41}[col]

        # DSW consumption: row 65 has adjusted consumption
        consumption = float(self._cell(ws, f"{col}65"))

        self.programs[key] = ProgramParams(
            name=f"DSW {col_map[col]}",
            pop_under5=float(self._cell(ws, f"{col}3")),
            pop_5_14=float(self._cell(ws, f"{col}42")),
            baseline_mortality_under5=float(self._cell(ws, f"{col}4")),
            baseline_mortality_over5=float(self._cell(ws, f"{col}17")),
            external_validity=float(self._cell(ev_ws, f"B{ev_row}")),
            plausibility_cap=float(self._cell(iv, f"B{cap_row}")),
            morbidity_ext_validity=float(self._cell(ev_ws, f"B{morb_ev_row}")),
            moral_weight_under5=float(self._cell(mw, f"B{mw_u5_row}")),
            moral_weight_over5=float(self._cell(mw, f"B{mw_o5_row}")),
            moral_weight_yld=float(self._cell(ws, f"{col}35")),
            adult_mortality_scaling=float(self._cell(am, f"{am_col}17")),
            ylds_per_100k=float(self._cell(ws, f"{col}29")),
            yld_under5=float(self._cell(ws, f"{col}43")),
            yld_5_14=float(self._cell(ws, f"{col}44")),
            dev_base=float(self._cell(ws, f"{col}39")),
            dev_chlorination_pct=float(self._cell(ws, f"{col}40")),
            baseline_diarrhea_incidence=float(self._cell(ws, f"{col}52")),
            cost_per_under5_diarrhea=float(self._cell(ws, f"{col}55")),
            consumption=consumption,
            iv_medical=float(self._cell(ws, f"{col}60")),
            ev_medical=float(self._cell(ws, f"{col}61")),
            value_doubling=float(self._cell(ws, f"{col}67")),
            mills_reincke=float(self._cell(ws, f"{col}58")),
            cost_per_person=float(self._cell(ws, f"{col}74")),
            cash_benchmark=float(self._cell(ws, f"{col}79")),
            # DSW adjustments
            downside_factor=float(self._cell(ws, f"{col}101")),
            excluded_effects_factor=float(self._cell(ws, f"{col}108")),
            # Leverage/funging
            has_leverage_adjustment=True,
            prob_govt_replace=float(self._cell(ws, f"{col}120")),
            prob_other_phil_replace=float(self._cell(ws, f"{col}121")),
            prob_upstream_stays=float(self._cell(ws, f"{col}122")),
            prob_not_exist=float(self._cell(ws, f"{col}123")),
            frac_govt_replace=float(self._cell(ws, f"{col}126")),
            frac_other_phil=float(self._cell(ws, f"{col}127")),
            frac_upstream=float(self._cell(ws, f"{col}128")),
            frac_not_exist=float(self._cell(ws, f"{col}129")),
            ce_govt=float(self._cell(ws, f"{col}133")),
            ce_other_phil=float(self._cell(ws, f"{col}134")),
            govt_costs=float(self._cell(ws, f"{col}112")),
            other_donor_upstream=float(self._cell(ws, f"{col}117")),
        )

    # ------------------------------------------------------------------
    # Formula chain
    # ------------------------------------------------------------------

    def compute_cost_effectiveness(
        self,
        program_key: str,
        **overrides: float,
    ) -> float:
        """Compute cost-effectiveness in multiples of GiveDirectly cash.

        Supported overrides:
            relative_risk, internal_validity_under5, internal_validity_over5,
            external_validity, plausibility_cap, internal_validity_morbidity,
            morbidity_ext_validity, mills_reincke
        """
        p = self.programs[program_key]

        rr = overrides.get("relative_risk", self.relative_risk)
        iv_u5 = overrides.get("internal_validity_under5", self.internal_validity_under5)
        iv_o5 = overrides.get("internal_validity_over5", self.internal_validity_over5)
        ev = overrides.get("external_validity", p.external_validity)
        cap = overrides.get("plausibility_cap", p.plausibility_cap)
        iv_morb = overrides.get("internal_validity_morbidity", self.internal_validity_morbidity)
        morb_ev = overrides.get("morbidity_ext_validity", p.morbidity_ext_validity)
        mr = overrides.get("mills_reincke", p.mills_reincke)

        # --- Under-5 mortality ---
        initial_estimate = (1 - rr) * iv_u5 * ev
        final_estimate = min(initial_estimate, cap)
        u5_deaths_per_100k = p.pop_under5 * p.baseline_mortality_under5 * final_estimate * 100_000
        u5_uv = u5_deaths_per_100k * p.moral_weight_under5

        # --- Over-5 mortality ---
        pop_over5 = 1 - p.pop_under5
        if initial_estimate > 0:
            cap_ratio = final_estimate / initial_estimate
        else:
            cap_ratio = 1.0
        o5_deaths_per_100k = (
            pop_over5
            * p.baseline_mortality_over5
            * (1 - rr)
            * p.adult_mortality_scaling
            * iv_o5
            * ev
            * cap_ratio
            * 100_000
        )
        o5_uv = o5_deaths_per_100k * p.moral_weight_over5

        # --- Morbidity ---
        morbidity_reduction = (1 - self.adjusted_diarrhea_rr) * iv_morb * morb_ev
        ylds_averted = p.ylds_per_100k * morbidity_reduction
        morb_uv = ylds_averted * p.moral_weight_yld

        # --- Development effects ---
        burden_ratio = p.yld_5_14 / p.yld_under5 if p.yld_under5 > 0 else 0.0
        raw_dev_uv = (
            (p.dev_base * p.dev_chlorination_pct * p.pop_under5)
            + (p.dev_base * p.dev_chlorination_pct * p.pop_5_14 * burden_ratio)
        ) * 100_000
        dev_uv = raw_dev_uv * iv_u5 * ev

        # --- Medical costs averted ---
        averted_cases = p.baseline_diarrhea_incidence * morbidity_reduction
        cost_per_u5 = averted_cases * (p.cost_per_under5_diarrhea / morbidity_reduction) if morbidity_reduction > 0 else 0.0
        # Actually: cost_per_u5 is read from spreadsheet directly as annual cost averted per under-5
        # The formula: B55 is pre-computed. B54 = B52 * B53. Then B55 is separate.
        # B57 = B55 * B56. Let me re-derive:
        # cost_per_person = cost_per_under5_diarrhea * pop_under5
        cost_per_person_served = p.cost_per_under5_diarrhea * p.pop_under5
        total_cost_averted = cost_per_person_served * mr
        adjusted_cost = total_cost_averted * p.iv_medical * p.ev_medical
        ln_consumption_increase = math.log(p.consumption + adjusted_cost) - math.log(p.consumption)
        value_per_ln_unit = p.value_doubling / math.log(2)
        uv_medical_per_person = ln_consumption_increase * value_per_ln_unit
        medical_uv = uv_medical_per_person * 100_000

        # --- Total units of value per 100k ---
        total_uv_per_100k = u5_uv + o5_uv + morb_uv + dev_uv + medical_uv

        # --- Cost-effectiveness ---
        uv_per_dollar = (total_uv_per_100k / 100_000) / p.cost_per_person
        uv_per_10k = uv_per_dollar * 10_000

        # --- DSW adjustments ---
        if p.has_leverage_adjustment:
            # Apply downside and excluded-effects adjustments
            adjusted_uv_per_10k = uv_per_10k * p.downside_factor * p.excluded_effects_factor

            # Leverage/funging: compute counterfactual value
            # Scenario UV values (what would happen in absence of DSW spending)
            cost_pp = p.cost_per_person  # annual cost per person served

            # Government replaces
            uv_govt = (adjusted_uv_per_10k * p.frac_govt_replace) - (
                p.ce_govt * cost_pp * p.frac_govt_replace * 10_000
            )

            # Other philanthropic actors replace
            uv_other_phil = (adjusted_uv_per_10k * p.frac_other_phil) - (
                p.ce_other_phil * cost_pp * p.frac_other_phil * 10_000
            )

            # Other actors' upstream costs stay the same
            uv_upstream = (adjusted_uv_per_10k * p.frac_upstream) + (
                (p.other_donor_upstream * p.ce_govt * 10_000) * (1 - p.frac_upstream)
            )

            # Program would not exist
            uv_not_exist = (adjusted_uv_per_10k * p.frac_not_exist) + (
                p.ce_govt * p.govt_costs * 10_000
            )

            # Counterfactual value (weighted by scenario probabilities)
            counterfactual = (
                p.prob_govt_replace * uv_govt
                + p.prob_other_phil_replace * uv_other_phil
                + p.prob_upstream_stays * uv_upstream
                + p.prob_not_exist * uv_not_exist
            )

            # Units of value attributed to DSW spending
            attributed_uv = adjusted_uv_per_10k - counterfactual
            ce_multiple = attributed_uv / p.cash_benchmark
        else:
            ce_multiple = uv_per_10k / p.cash_benchmark

        return ce_multiple

    def detect_cap_binding(self, program_key: str, **overrides: float) -> bool:
        """Return True if the plausibility cap binds for this program."""
        p = self.programs[program_key]

        rr = overrides.get("relative_risk", self.relative_risk)
        iv_u5 = overrides.get("internal_validity_under5", self.internal_validity_under5)
        ev = overrides.get("external_validity", p.external_validity)
        cap = overrides.get("plausibility_cap", p.plausibility_cap)

        initial_estimate = (1 - rr) * iv_u5 * ev
        return initial_estimate > cap

    def run_sensitivity(
        self,
        program_key: str,
        parameter_name: str,
        low: float,
        central: float,
        high: float,
    ) -> dict[str, Any]:
        """Run sensitivity analysis for a single parameter.

        Returns dict with baseline CE, CE at low/central/high,
        percentage changes, and cap binding info.
        """
        baseline_ce = self.compute_cost_effectiveness(program_key)
        cap_binds_baseline = self.detect_cap_binding(program_key)

        results: dict[str, float | bool | str] = {
            "parameter_name": parameter_name,
            "baseline": baseline_ce,
            "cap_binds_baseline": cap_binds_baseline,
        }

        for label, value in [("low", low), ("central", central), ("high", high)]:
            ce = self.compute_cost_effectiveness(program_key, **{parameter_name: value})
            cap_binds = self.detect_cap_binding(program_key, **{parameter_name: value})
            pct_change = ((ce - baseline_ce) / baseline_ce) * 100 if baseline_ce != 0 else 0.0
            results[label] = ce
            results[f"pct_change_{label}"] = pct_change
            results[f"cap_binds_{label}"] = cap_binds

        return results

    def get_parameter_summary(self) -> str:
        """Return a human-readable markdown summary of key parameters and baseline CE."""
        lines: list[str] = []
        lines.append("# Water Chlorination CEA — Parameter Summary\n")

        lines.append("## Shared Parameters\n")
        lines.append(f"- **Pooled ln(RR):** {self.pooled_ln_rr:.10f}")
        lines.append(f"- **Relative risk of all-cause mortality:** {self.relative_risk:.10f}")
        lines.append(f"- **Internal validity, under-5 mortality:** {self.internal_validity_under5:.10f}")
        lines.append(f"- **Internal validity, over-5 mortality:** {self.internal_validity_over5:.10f}")
        lines.append(f"- **Internal validity, morbidity:** {self.internal_validity_morbidity:.4f}")
        lines.append(f"- **Adjusted diarrhea RR:** {self.adjusted_diarrhea_rr:.4f}")
        lines.append(f"- **Mills-Reincke multiplier:** {self.mills_reincke:.10f}")
        lines.append("")

        lines.append("## Program-Specific Parameters\n")
        for key in self.PROGRAMS:
            p = self.programs[key]
            ce = self.compute_cost_effectiveness(key)
            cap_binds = self.detect_cap_binding(key)
            lines.append(f"### {p.name}\n")
            lines.append(f"- **Cost-effectiveness (x cash):** {ce:.4f}")
            lines.append(f"- **External validity:** {p.external_validity:.10f}")
            lines.append(f"- **Plausibility cap:** {p.plausibility_cap:.4f} (binds: {cap_binds})")
            lines.append(f"- **Pop under-5:** {p.pop_under5:.4f}")
            lines.append(f"- **Baseline mortality under-5:** {p.baseline_mortality_under5:.10f}")
            lines.append(f"- **Baseline mortality over-5:** {p.baseline_mortality_over5:.10f}")
            lines.append(f"- **Adult mortality scaling:** {p.adult_mortality_scaling:.10f}")
            lines.append(f"- **Moral weight under-5:** {p.moral_weight_under5:.4f}")
            lines.append(f"- **Moral weight over-5:** {p.moral_weight_over5:.4f}")
            lines.append(f"- **Cost per person:** {p.cost_per_person:.10f}")
            lines.append(f"- **Consumption:** {p.consumption:.4f}")
            lines.append("")

        return "\n".join(lines)


# ---------------------------------------------------------------------------
# ITN (Insecticide-Treated Nets) CEA
# ---------------------------------------------------------------------------


@dataclass
class ITNProgramParams:
    """Country/program-specific parameters for ITN CEA."""

    name: str
    # Coverage (person-years from Main CEA rows 55-57)
    person_years_u5: float
    person_years_5_14: float
    person_years_over14: float
    # Mortality (rows 70, 73-75, 80-81)
    direct_malaria_mortality_u5: float
    smc_reduction: float
    baseline_net_coverage: float
    malaria_deaths_u5_national: float
    total_malaria_deaths_national: float
    # Incidence (rows 88-89, from Counterfactual malaria sheet)
    incidence_u5_no_nets: float
    incidence_5_14_no_nets: float
    # Per-country resistance (row 66)
    insecticide_resistance: float
    # Final adjustments (Simple CEA rows 32-35)
    additional_benefits_adj: float
    grantee_adj: float
    leverage_adj: float
    funging_adj: float
    # Grant (row 8)
    grant_size: float


class ITNCEA:
    """Reads GiveWell ITN (insecticide-treated nets) CEA and replicates the formula chain.

    Columnar layout: 8 programs across columns I-P of InsecticideCEA.xlsx.
    """

    PROGRAMS = (
        "chad", "drc", "guinea", "nigeria_gf",
        "nigeria_pmi", "south_sudan", "togo", "uganda",
    )

    _COL_MAP: dict[str, str] = {
        "chad": "I", "drc": "J", "guinea": "K", "nigeria_gf": "L",
        "nigeria_pmi": "M", "south_sudan": "N", "togo": "O", "uganda": "P",
    }

    _NAMES: dict[str, str] = {
        "chad": "Chad", "drc": "DRC", "guinea": "Guinea",
        "nigeria_gf": "Nigeria (GF)", "nigeria_pmi": "Nigeria (PMI)",
        "south_sudan": "South Sudan", "togo": "Togo", "uganda": "Uganda",
    }

    def __init__(self, path: Path) -> None:
        self._path = Path(path)
        wb = openpyxl.load_workbook(str(self._path), data_only=True)
        self._read_shared_params(wb)
        self._read_program_params(wb)
        wb.close()

    # ------------------------------------------------------------------
    # Reading helpers
    # ------------------------------------------------------------------

    def _cell(self, ws: Any, ref: str) -> Any:
        val = ws[ref].value
        if val is None:
            raise ValueError(f"Cell {ws.title}!{ref} is None")
        return val

    def _read_shared_params(self, wb: openpyxl.Workbook) -> None:
        ws = wb["Main CEA"]
        self.incidence_reduction: float = float(self._cell(ws, "H62"))
        self.net_usage_trial: float = float(self._cell(ws, "H35"))
        self.internal_validity: float = float(self._cell(ws, "H64"))
        self.external_validity: float = float(self._cell(ws, "H65"))
        self.mortality_incidence_ratio: float = float(self._cell(ws, "H68"))
        self.indirect_deaths_multiplier: float = float(self._cell(ws, "H71"))
        self.over5_relative_efficacy: float = float(self._cell(ws, "H84"))
        self.income_per_case: float = float(self._cell(ws, "H96"))
        self.years_to_benefits: int = int(self._cell(ws, "H98"))
        self.discount_rate: float = float(self._cell(ws, "H99"))
        self.benefit_duration: int = int(self._cell(ws, "H101"))
        self.household_multiplier: float = float(self._cell(ws, "H103"))
        self.moral_weight_u5: float = float(self._cell(ws, "H112"))
        self.moral_weight_over5: float = float(self._cell(ws, "H116"))
        self.ln_consumption_value: float = float(self._cell(ws, "H120"))
        self.benchmark: float = float(self._cell(ws, "H144"))

    def _read_program_params(self, wb: openpyxl.Workbook) -> None:
        ws = wb["Main CEA"]
        sc = wb["Simple CEA"]
        self.programs: dict[str, ITNProgramParams] = {}
        for key, col in self._COL_MAP.items():
            self.programs[key] = ITNProgramParams(
                name=self._NAMES[key],
                person_years_u5=float(self._cell(ws, f"{col}55")),
                person_years_5_14=float(self._cell(ws, f"{col}56")),
                person_years_over14=float(self._cell(ws, f"{col}57")),
                direct_malaria_mortality_u5=float(self._cell(ws, f"{col}70")),
                smc_reduction=float(self._cell(ws, f"{col}73")),
                baseline_net_coverage=float(self._cell(ws, f"{col}75")),
                malaria_deaths_u5_national=float(self._cell(ws, f"{col}80")),
                total_malaria_deaths_national=float(self._cell(ws, f"{col}81")),
                incidence_u5_no_nets=float(self._cell(ws, f"{col}88")),
                incidence_5_14_no_nets=float(self._cell(ws, f"{col}89")),
                insecticide_resistance=float(self._cell(ws, f"{col}66")),
                additional_benefits_adj=float(self._cell(sc, f"{col}32")),
                grantee_adj=float(self._cell(sc, f"{col}33")),
                leverage_adj=float(self._cell(sc, f"{col}34")),
                funging_adj=float(self._cell(sc, f"{col}35")),
                grant_size=float(self._cell(ws, "H8")),
            )

    # ------------------------------------------------------------------
    # Formula chain
    # ------------------------------------------------------------------

    def _expected_reduction(
        self,
        incidence_reduction: float,
        internal_validity: float,
        external_validity: float,
        insecticide_resistance: float,
    ) -> float:
        """Compute expected reduction in malaria incidence/mortality for net sleepers."""
        implied = incidence_reduction / self.net_usage_trial
        return (
            implied
            * (1 + internal_validity)
            * (1 + external_validity)
            * (1 + insecticide_resistance)
        )

    def _annuity_due_factor(self) -> float:
        """Present value of annuity-due (payments at start of period)."""
        r = self.discount_rate
        n = self.benefit_duration
        ordinary = (1 - (1 + r) ** (-n)) / r
        return ordinary * (1 + r)

    def compute_cost_effectiveness(
        self,
        program_key: str,
        **overrides: float,
    ) -> float:
        """Compute cost-effectiveness in multiples of GiveDirectly cash.

        Supported overrides:
            incidence_reduction, internal_validity, external_validity,
            insecticide_resistance, indirect_deaths_multiplier,
            moral_weight_under5, moral_weight_over5
        """
        p = self.programs[program_key]

        ir = overrides.get("incidence_reduction", self.incidence_reduction)
        iv = overrides.get("internal_validity", self.internal_validity)
        ev = overrides.get("external_validity", self.external_validity)
        resist = overrides.get("insecticide_resistance", p.insecticide_resistance)
        indirect = overrides.get("indirect_deaths_multiplier", self.indirect_deaths_multiplier)
        mw_u5 = overrides.get("moral_weight_under5", self.moral_weight_u5)
        mw_o5 = overrides.get("moral_weight_over5", self.moral_weight_over5)

        # --- Expected reduction ---
        exp_red = self._expected_reduction(ir, iv, ev, resist)
        exp_mort_red = exp_red * self.mortality_incidence_ratio

        # --- Malaria mortality in absence of nets ---
        total_mort = p.direct_malaria_mortality_u5 * (1 + indirect)
        adj_mort = total_mort - p.smc_reduction
        denom = 1 - p.baseline_net_coverage * exp_mort_red
        if denom <= 0:
            denom = 1e-9  # avoid division by zero for extreme parameters
        mort_no_nets = adj_mort / denom

        # --- Under-5 deaths averted ---
        deaths_u5 = p.person_years_u5 * mort_no_nets * exp_mort_red

        # --- Over-5 deaths averted ---
        over5_deaths_national = (
            p.total_malaria_deaths_national - p.malaria_deaths_u5_national
        )
        if p.malaria_deaths_u5_national > 0:
            over5_ratio = over5_deaths_national / p.malaria_deaths_u5_national
        else:
            over5_ratio = 0.0
        deaths_over5 = deaths_u5 * over5_ratio * self.over5_relative_efficacy

        # --- Development benefits (long-term income) ---
        cases_u5 = p.person_years_u5 * p.incidence_u5_no_nets * exp_red
        cases_5_14 = p.person_years_5_14 * p.incidence_5_14_no_nets * exp_red
        total_cases = cases_u5 + cases_5_14

        adj_income = math.log(1 + self.income_per_case)
        discounted = adj_income * (1 / (1 + self.discount_rate)) ** self.years_to_benefits
        pv_per_case = discounted * self._annuity_due_factor() * self.household_multiplier
        dev_value = total_cases * pv_per_case * self.ln_consumption_value

        # --- Total value ---
        u5_value = deaths_u5 * mw_u5
        over5_value = deaths_over5 * mw_o5
        total_value = u5_value + over5_value + dev_value

        # --- Initial CE (before final adjustments) ---
        uv_per_dollar = total_value / p.grant_size
        initial_ce = uv_per_dollar / self.benchmark

        # --- Final CE (with project-level adjustments) ---
        final_ce = (
            initial_ce
            * (1 + p.additional_benefits_adj)
            * (1 + p.grantee_adj)
            * (1 + p.leverage_adj + p.funging_adj)
        )

        return final_ce

    def detect_cap_binding(self, program_key: str, **overrides: float) -> bool:
        """ITN CEA does not use plausibility caps."""
        return False

    def run_sensitivity(
        self,
        program_key: str,
        parameter_name: str,
        low: float,
        central: float,
        high: float,
    ) -> dict[str, Any]:
        """Run sensitivity analysis for a single parameter.

        Returns dict with baseline CE, CE at low/central/high,
        percentage changes, and cap binding info.
        """
        baseline_ce = self.compute_cost_effectiveness(program_key)

        results: dict[str, float | bool | str] = {
            "parameter_name": parameter_name,
            "baseline": baseline_ce,
            "cap_binds_baseline": False,
        }

        for label, value in [("low", low), ("central", central), ("high", high)]:
            ce = self.compute_cost_effectiveness(program_key, **{parameter_name: value})
            pct_change = ((ce - baseline_ce) / baseline_ce) * 100 if baseline_ce != 0 else 0.0
            results[label] = ce
            results[f"pct_change_{label}"] = pct_change
            results[f"cap_binds_{label}"] = False

        return results

    def get_parameter_summary(self) -> str:
        """Return a human-readable markdown summary of key parameters and baseline CE."""
        lines: list[str] = []
        lines.append("# Insecticide-Treated Nets CEA — Parameter Summary\n")

        lines.append("## Shared Parameters\n")
        lines.append(f"- **Malaria incidence reduction (Pryce et al.):** {self.incidence_reduction}")
        lines.append(f"- **Net usage in trials:** {self.net_usage_trial}")
        lines.append(f"- **Internal validity adjustment:** {self.internal_validity}")
        lines.append(f"- **External validity adjustment:** {self.external_validity}")
        lines.append(f"- **Indirect deaths per direct death:** {self.indirect_deaths_multiplier}")
        lines.append(f"- **Over-5 relative efficacy:** {self.over5_relative_efficacy}")
        lines.append(f"- **Moral weight under-5:** {self.moral_weight_u5:.4f}")
        lines.append(f"- **Moral weight over-5:** {self.moral_weight_over5:.4f}")
        lines.append(f"- **Income per case averted:** {self.income_per_case}")
        lines.append(f"- **Discount rate:** {self.discount_rate}")
        lines.append(f"- **Benchmark (UoV/$):** {self.benchmark}")
        lines.append("")

        lines.append("## Program-Specific Parameters\n")
        for key in self.PROGRAMS:
            p = self.programs[key]
            ce = self.compute_cost_effectiveness(key)
            lines.append(f"### {p.name}\n")
            lines.append(f"- **Cost-effectiveness (x cash):** {ce:.2f}")
            lines.append(f"- **Insecticide resistance:** {p.insecticide_resistance:.4f}")
            lines.append(f"- **Direct malaria mortality (u5):** {p.direct_malaria_mortality_u5:.6f}")
            lines.append(f"- **SMC reduction:** {p.smc_reduction:.6f}")
            lines.append(f"- **Baseline net coverage:** {p.baseline_net_coverage:.3f}")
            lines.append(f"- **Person-years u5:** {p.person_years_u5:.0f}")
            lines.append(f"- **Person-years 5-14:** {p.person_years_5_14:.0f}")
            lines.append(f"- **Additional benefits adj:** {p.additional_benefits_adj}")
            lines.append(f"- **Leverage + funging adj:** {p.leverage_adj + p.funging_adj:.4f}")
            lines.append("")

        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Malaria CEA (SMC cost-per-cycle model)
# ---------------------------------------------------------------------------

@dataclass
class SMCCountryParams:
    """Country-specific parameters for the SMC cost model."""

    name: str
    total_spending_mc: float       # Malaria Consortium spending
    total_spending_other: float    # Other philanthropic actors
    total_spending_govt: float     # Government spending
    total_spending: float          # Sum of all
    target_population: float       # Reported target population
    adjusted_person_months: float  # Person-months of coverage (with adjustments)
    cost_per_cycle: float          # Pre-computed cost per cycle
    cost_per_child_4: float        # Cost per child (4 cycles)
    cost_per_child_5: float        # Cost per child (5 cycles)


class MalariaCEA:
    """Reads GiveWell Malaria CEA (SMC cost model) and runs sensitivity analysis.

    The spreadsheet calculates cost per SMC cycle administered. Key adjustable
    parameters are self-report bias and adherence adjustment, which affect the
    denominator (person-months of coverage).

    Formula chain:
        raw_person_months = adjusted_person_months / (self_report_bias * adherence)
        person_months = raw_person_months * self_report_bias * adherence
        cost_per_cycle = total_costs / person_months
        cost_per_child_4 = cost_per_cycle * 4
    """

    PROGRAMS = ("burkina_faso", "chad", "nigeria", "togo")

    # Column mapping: country -> spreadsheet column in Per-country analysis
    _COL_MAP = {
        "burkina_faso": "B",
        "chad": "C",
        "nigeria": "D",
        "togo": "E",
    }

    def __init__(self, path: Path) -> None:
        self._path = Path(path)
        wb = openpyxl.load_workbook(str(self._path), data_only=True)
        self._read_params(wb)
        wb.close()

    def _cell(self, ws: Any, ref: str) -> Any:
        """Read a cell value, raising if None."""
        val = ws[ref].value
        if val is None:
            raise ValueError(f"Cell {ws.title}!{ref} is None")
        return val

    def _read_params(self, wb: openpyxl.Workbook) -> None:
        pca = wb["Per-country analysis (2018-2023"]

        # Shared adjustment parameters
        self.self_report_bias: float = float(self._cell(pca, "B16"))
        self.adherence_adjustment: float = float(self._cell(pca, "B17"))

        # Adherence sub-parameters
        adh = wb["Calc Adherence adjustment"]
        self.social_desirability_bias: float = float(self._cell(adh, "B5"))
        self.efficacy_reduction: float = float(self._cell(adh, "B6"))

        # Per-country adherence adjustments
        self.country_adherence: dict[str, float] = {
            "burkina_faso": float(self._cell(adh, "B32")),
            "chad": float(self._cell(adh, "C32")),
            "nigeria": float(self._cell(adh, "D32")),
            "togo": self.adherence_adjustment,  # Togo uses overall average
        }

        # Per-country weighted average adherence rates (before social desirability)
        self.country_raw_adherence: dict[str, float] = {
            "burkina_faso": float(self._cell(adh, "B30")),
            "chad": float(self._cell(adh, "C30")),
            "nigeria": float(self._cell(adh, "D30")),
            "togo": self.adherence_adjustment,
        }

        # Per-country parameters
        self.programs: dict[str, SMCCountryParams] = {}
        for key, col in self._COL_MAP.items():
            name = {
                "burkina_faso": "Burkina Faso",
                "chad": "Chad",
                "nigeria": "Nigeria",
                "togo": "Togo",
            }[key]
            self.programs[key] = SMCCountryParams(
                name=name,
                total_spending_mc=float(self._cell(pca, f"{col}5")),
                total_spending_other=float(self._cell(pca, f"{col}6")),
                total_spending_govt=float(self._cell(pca, f"{col}7")),
                total_spending=float(self._cell(pca, f"{col}8")),
                target_population=float(self._cell(pca, f"{col}9")),
                adjusted_person_months=float(self._cell(pca, f"{col}20")),
                cost_per_cycle=float(self._cell(pca, f"{col}23")),
                cost_per_child_4=float(self._cell(pca, f"{col}24")),
                cost_per_child_5=float(self._cell(pca, f"{col}25")),
            )

        # Back-calculate raw (unadjusted) person-months per country
        adjustment_product = self.self_report_bias * self.adherence_adjustment
        self._raw_person_months: dict[str, float] = {}
        for key, p in self.programs.items():
            self._raw_person_months[key] = p.adjusted_person_months / adjustment_product

    # ------------------------------------------------------------------
    # Formula chain
    # ------------------------------------------------------------------

    def compute_cost_effectiveness(
        self,
        program_key: str,
        **overrides: float,
    ) -> float:
        """Compute cost per child treated with 4 cycles.

        Supported overrides:
            self_report_bias, adherence_adjustment,
            social_desirability_bias, efficacy_reduction,
            govt_cost_fraction, total_spending
        """
        p = self.programs[program_key]

        sr = overrides.get("self_report_bias", self.self_report_bias)
        adh = overrides.get("adherence_adjustment", self.adherence_adjustment)

        # If social_desirability_bias or efficacy_reduction are overridden,
        # recompute adherence from the sub-parameters
        if "social_desirability_bias" in overrides or "efficacy_reduction" in overrides:
            sdb = overrides.get("social_desirability_bias", self.social_desirability_bias)
            eff_red = overrides.get("efficacy_reduction", self.efficacy_reduction)
            raw_adh = self.country_raw_adherence.get(program_key, 0.97)
            adjusted_adh = raw_adh * sdb
            adh = 1 - (1 - adjusted_adh) * eff_red

        total_costs = overrides.get("total_spending", p.total_spending)

        # Recompute person-months with overridden adjustments
        raw_pm = self._raw_person_months[program_key]
        person_months = raw_pm * sr * adh

        if person_months <= 0:
            return float("inf")

        cost_per_cycle = total_costs / person_months
        cost_per_child_4 = cost_per_cycle * 4
        return cost_per_child_4

    def detect_cap_binding(self, program_key: str, **overrides: float) -> bool:
        """No plausibility cap concept in the SMC cost model."""
        return False

    def run_sensitivity(
        self,
        program_key: str,
        parameter_name: str,
        low: float,
        central: float,
        high: float,
    ) -> dict[str, Any]:
        """Run sensitivity analysis for a single parameter."""
        baseline_ce = self.compute_cost_effectiveness(program_key)

        results: dict[str, float | bool | str] = {
            "parameter_name": parameter_name,
            "baseline": baseline_ce,
            "cap_binds_baseline": False,
        }

        for label, value in [("low", low), ("central", central), ("high", high)]:
            ce = self.compute_cost_effectiveness(program_key, **{parameter_name: value})
            pct_change = ((ce - baseline_ce) / baseline_ce) * 100 if baseline_ce != 0 else 0.0
            results[label] = ce
            results[f"pct_change_{label}"] = pct_change
            results[f"cap_binds_{label}"] = False

        return results

    def get_parameter_summary(self) -> str:
        """Return a human-readable markdown summary of key parameters and baseline costs."""
        lines: list[str] = []
        lines.append("# SMC (Seasonal Malaria Chemoprevention) CEA — Parameter Summary\n")

        lines.append("## Shared Adjustment Parameters\n")
        lines.append(f"- **Self-report bias adjustment:** {self.self_report_bias:.4f}")
        lines.append(f"- **Adherence adjustment (weighted avg):** {self.adherence_adjustment:.10f}")
        lines.append(f"- **Social desirability bias:** {self.social_desirability_bias:.4f}")
        lines.append(f"- **Efficacy reduction for non-adherence:** {self.efficacy_reduction:.4f}")
        lines.append("")

        lines.append("## Per-Country Parameters\n")
        for key in self.PROGRAMS:
            p = self.programs[key]
            ce = self.compute_cost_effectiveness(key)
            lines.append(f"### {p.name}\n")
            lines.append(f"- **Cost per child treated (4 cycles):** ${ce:.4f}")
            lines.append(f"- **Cost per SMC cycle:** ${p.cost_per_cycle:.4f}")
            lines.append(f"- **Total spending:** ${p.total_spending:,.2f}")
            lines.append(f"  - Malaria Consortium: ${p.total_spending_mc:,.2f}")
            lines.append(f"  - Other philanthropic: ${p.total_spending_other:,.2f}")
            lines.append(f"  - Government: ${p.total_spending_govt:,.2f}")
            lines.append(f"- **Target population (2018-2023):** {p.target_population:,.0f}")
            lines.append(
                f"- **Adjusted person-months of coverage:** {p.adjusted_person_months:,.2f}"
            )
            adh = self.country_adherence.get(key, self.adherence_adjustment)
            lines.append(f"- **Country adherence adjustment:** {adh:.10f}")
            lines.append("")

        return "\n".join(lines)


# ---------------------------------------------------------------------------
# VAS (Vitamin A Supplementation) CEA
# ---------------------------------------------------------------------------


@dataclass
class VASLocationParams:
    """Per-location parameters for VAS CEA."""

    name: str
    grantee: str  # "Helen Keller International" or "Nutrition International"

    # Final CE (from Simple CEA row 37)
    ce_multiple: float

    # Initial CE before adjustments (Simple CEA row 21)
    initial_ce_multiple: float

    # Cost (Inputs sheet)
    cost_per_supplement: float  # row 18
    pct_costs_grantee: float  # row 11
    pct_costs_ni: float  # row 12 (Nutrition International capsule procurement)
    pct_costs_govt_financial: float  # row 14
    pct_costs_govt_inkind: float  # row 15

    # Mortality (Counterfactual mortality sheet)
    total_under5_deaths: float  # row 8
    early_neonatal_deaths: float  # row 9
    late_neonatal_deaths: float  # row 10
    post_neonatal_deaths: float  # row 11
    mortality_rate_6_59mo: float  # row 24 (annual rate among 6-59mo)

    # External validity
    vad_prevalence: float  # External validity row 8
    vad_survey_year: int  # External validity row 9
    estimated_vad_2021: float  # External validity row 46
    pct_under5_6_59mo: float  # Counterfactual mortality row 22

    # Adjustments (Simple CEA rows 31-34)
    adj_additional_benefits: float  # row 31
    adj_grantee_factors: float  # row 32
    adj_leverage: float  # row 33
    adj_funging: float  # row 34

    # Nigeria aggregation (only populated for the "nigeria" aggregate entry)
    ce_multiple_min: float = 0.0
    ce_multiple_max: float = 0.0
    ce_multiple_mean: float = 0.0
    n_states: int = 0


class VASCEA:
    """Reads GiveWell VAS CEA and produces a rich parameter summary.

    Unlike WaterCEA/ITNCEA/MalariaCEA, this class reads pre-computed CE
    multiples rather than replicating the formula chain. This is a deliberate
    trade-off: the VAS model spans 5+ sheets with 37 location columns and
    cross-sheet DUMMYFUNCTION references, making replication fragile and
    unnecessary — the pipeline agents consume get_parameter_summary() text,
    not Python computations.

    Consequence: compute_cost_effectiveness() raises NotImplementedError if
    overrides are passed. Any future Bayesian optimization on VAS parameters
    would require a full formula-chain implementation.

    Location layout (Inputs/Simple CEA sheets):
        Col I-P: HKI countries (Burkina Faso through Niger)
        Col Q-AK: 21 Nigerian locations (20 states + FCT), aggregated as one entry
        Col AL-AO: Nutrition International countries (Angola, Chad, Togo, Uganda)
    """

    PROGRAMS = (
        "burkina_faso", "cameroon", "cote_divoire", "drc", "guinea",
        "madagascar", "mali", "niger", "angola", "chad", "togo", "uganda",
        "nigeria",
    )

    # Non-Nigerian locations: key -> (column index in Inputs/Simple CEA, grantee)
    _LOC_COLS: dict[str, tuple[int, str]] = {
        "burkina_faso": (9, "Helen Keller International"),
        "cameroon": (10, "Helen Keller International"),
        "cote_divoire": (11, "Helen Keller International"),
        "drc": (12, "Helen Keller International"),
        "guinea": (13, "Helen Keller International"),
        "madagascar": (14, "Helen Keller International"),
        "mali": (15, "Helen Keller International"),
        "niger": (16, "Helen Keller International"),
        "angola": (38, "Nutrition International"),
        "chad": (39, "Nutrition International"),
        "togo": (40, "Nutrition International"),
        "uganda": (41, "Nutrition International"),
    }

    _LOC_NAMES: dict[str, str] = {
        "burkina_faso": "Burkina Faso",
        "cameroon": "Cameroon",
        "cote_divoire": "Cote d'Ivoire",
        "drc": "DRC",
        "guinea": "Guinea",
        "madagascar": "Madagascar",
        "mali": "Mali",
        "niger": "Niger",
        "angola": "Angola",
        "chad": "Chad",
        "togo": "Togo",
        "uganda": "Uganda",
    }

    # Nigerian location columns: Q=17 through AK=37 (20 states + FCT)
    _NIGERIA_COLS: dict[str, int] = {
        "Adamawa": 17, "Akwa Ibom": 18, "Anambra": 19, "Benue": 20,
        "Delta": 21, "Ebonyi": 22, "Edo": 23, "Ekiti": 24,
        "Imo": 25, "Kogi": 26, "Nasarawa": 27, "Ogun": 28,
        "Osun": 29, "Rivers": 30, "Taraba": 31, "Kaduna": 32,
        "Niger": 33, "Plateau": 34, "Sokoto": 35, "Kebbi": 36,
        "FCT (Abuja)": 37,
    }

    # Sensitivity analysis: location -> start column (3 cols each: 25th, best, 75th)
    _SENS_COLS: dict[str, int] = {
        "burkina_faso": 6, "cameroon": 9, "cote_divoire": 12,
        "drc": 15, "guinea": 18, "madagascar": 21,
        "mali": 24, "niger": 27,
        "angola": 93, "chad": 96, "togo": 99, "uganda": 102,
    }

    # Sensitivity parameter rows (percentage-change section, rows 56-67)
    _SENS_ROWS: dict[str, int] = {
        "cost_per_person": 56,
        "counterfactual_coverage": 57,
        "mortality_rate": 58,
        "effect_of_vas_on_mortality": 59,
        "developmental_benefits": 62,
        "additional_benefits_downsides": 65,
        "grantee_factors": 66,
        "funging": 67,
    }

    def __init__(self, path: Path) -> None:
        self._path = Path(path)
        wb = openpyxl.load_workbook(str(self._path), data_only=True)
        self._read_shared(wb)
        self._read_locations(wb)
        self._read_nigeria_aggregate(wb)
        self._read_sensitivity(wb)
        wb.close()

    def _safe_float(self, ws: Any, row: int, col: int, default: float = 0.0) -> float:
        """Read a cell as float, returning default for None or non-numeric values."""
        val = ws.cell(row=row, column=col).value
        if val is None or val == "-" or val == "":
            return default
        try:
            return float(val)
        except (ValueError, TypeError):
            return default

    def _safe_int(self, ws: Any, row: int, col: int, default: int = 0) -> int:
        """Read a cell as int, returning default for None or non-numeric values."""
        val = ws.cell(row=row, column=col).value
        if val is None or val == "-" or val == "" or val == "n/a":
            return default
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return default

    def _read_shared(self, wb: openpyxl.Workbook) -> None:
        """Read shared (non-location-specific) parameters."""
        inp = wb["Inputs"]
        sc = wb["Simple CEA"]
        cm = wb["Counterfactual mortality"]

        self.grant_size: float = float(inp.cell(row=8, column=8).value)  # H8
        self.rounds_per_year: float = float(inp.cell(row=19, column=8).value)  # H19
        self.moral_value_under5: float = float(sc.cell(row=20, column=8).value)  # Simple CEA H20
        self.prop_postneonatal_1_5mo: float = float(cm.cell(row=14, column=8).value)  # CM H14

    def _read_locations(self, wb: openpyxl.Workbook) -> None:
        """Read per-location parameters for all non-Nigerian locations."""
        inp = wb["Inputs"]
        sc = wb["Simple CEA"]
        cm = wb["Counterfactual mortality"]
        ev = wb["External validity"]

        self.locations: dict[str, VASLocationParams] = {}

        for key, (col, grantee) in self._LOC_COLS.items():
            self.locations[key] = VASLocationParams(
                name=self._LOC_NAMES[key],
                grantee=grantee,
                ce_multiple=self._safe_float(sc, 37, col),
                initial_ce_multiple=self._safe_float(sc, 21, col),
                cost_per_supplement=self._safe_float(inp, 18, col),
                pct_costs_grantee=self._safe_float(inp, 11, col),
                pct_costs_ni=self._safe_float(inp, 12, col),
                pct_costs_govt_financial=self._safe_float(inp, 14, col),
                pct_costs_govt_inkind=self._safe_float(inp, 15, col),
                total_under5_deaths=self._safe_float(cm, 8, col),
                early_neonatal_deaths=self._safe_float(cm, 9, col),
                late_neonatal_deaths=self._safe_float(cm, 10, col),
                post_neonatal_deaths=self._safe_float(cm, 11, col),
                mortality_rate_6_59mo=self._safe_float(cm, 24, col),
                vad_prevalence=self._safe_float(ev, 8, col),
                vad_survey_year=self._safe_int(ev, 9, col),
                estimated_vad_2021=self._safe_float(ev, 46, col),
                pct_under5_6_59mo=self._safe_float(cm, 22, col),
                adj_additional_benefits=self._safe_float(sc, 31, col),
                adj_grantee_factors=self._safe_float(sc, 32, col),
                adj_leverage=self._safe_float(sc, 33, col),
                adj_funging=self._safe_float(sc, 34, col),
            )

    def _read_nigeria_aggregate(self, wb: openpyxl.Workbook) -> None:
        """Aggregate 21 Nigerian locations (20 states + FCT) into a single summary entry."""
        sc = wb["Simple CEA"]
        inp = wb["Inputs"]
        cm = wb["Counterfactual mortality"]
        ev = wb["External validity"]

        state_final_ce: list[float] = []
        state_initial_ce: list[float] = []
        state_costs: list[float] = []
        total_deaths = 0.0
        total_early_neo = 0.0
        total_late_neo = 0.0
        total_post_neo = 0.0

        self.nigeria_states: dict[str, float] = {}

        for state_name, col in self._NIGERIA_COLS.items():
            final_ce = self._safe_float(sc, 37, col)
            initial_ce = self._safe_float(sc, 21, col)
            if final_ce > 0:
                state_final_ce.append(final_ce)
                self.nigeria_states[state_name] = final_ce
            if initial_ce > 0:
                state_initial_ce.append(initial_ce)
            cost = self._safe_float(inp, 18, col)
            if cost > 0:
                state_costs.append(cost)
            total_deaths += self._safe_float(cm, 8, col)
            total_early_neo += self._safe_float(cm, 9, col)
            total_late_neo += self._safe_float(cm, 10, col)
            total_post_neo += self._safe_float(cm, 11, col)

        first_col = 17
        mean_final = sum(state_final_ce) / len(state_final_ce) if state_final_ce else 0.0
        mean_initial = sum(state_initial_ce) / len(state_initial_ce) if state_initial_ce else 0.0

        self.locations["nigeria"] = VASLocationParams(
            name="Nigeria (20 states + FCT)",
            grantee="Helen Keller International",
            ce_multiple=mean_final,
            initial_ce_multiple=mean_initial,
            cost_per_supplement=sum(state_costs) / len(state_costs) if state_costs else 0.0,
            pct_costs_grantee=self._safe_float(inp, 11, first_col),
            pct_costs_ni=self._safe_float(inp, 12, first_col),
            pct_costs_govt_financial=self._safe_float(inp, 14, first_col),
            pct_costs_govt_inkind=self._safe_float(inp, 15, first_col),
            total_under5_deaths=total_deaths,
            early_neonatal_deaths=total_early_neo,
            late_neonatal_deaths=total_late_neo,
            post_neonatal_deaths=total_post_neo,
            mortality_rate_6_59mo=self._safe_float(cm, 24, first_col),
            vad_prevalence=self._safe_float(ev, 8, first_col),
            vad_survey_year=self._safe_int(ev, 9, first_col),
            estimated_vad_2021=self._safe_float(ev, 46, first_col),
            pct_under5_6_59mo=self._safe_float(cm, 22, first_col),
            adj_additional_benefits=self._safe_float(sc, 31, first_col),
            adj_grantee_factors=self._safe_float(sc, 32, first_col),
            adj_leverage=self._safe_float(sc, 33, first_col),
            adj_funging=self._safe_float(sc, 34, first_col),
            ce_multiple_min=min(state_final_ce) if state_final_ce else 0.0,
            ce_multiple_max=max(state_final_ce) if state_final_ce else 0.0,
            ce_multiple_mean=mean_final,
            n_states=len(state_final_ce),
        )

    def _read_sensitivity(self, wb: openpyxl.Workbook) -> None:
        """Read sensitivity percentage changes from the Sensitivity analysis sheet."""
        sa = wb["Sensitivity analysis"]

        self.sensitivity: dict[str, dict[str, dict[str, float]]] = {}
        for loc_key, start_col in self._SENS_COLS.items():
            loc_sens: dict[str, dict[str, float]] = {}
            for param_name, row in self._SENS_ROWS.items():
                pct_25th = self._safe_float(sa, row, start_col)
                pct_75th = self._safe_float(sa, row, start_col + 2)
                loc_sens[param_name] = {
                    "pct_change_25th": pct_25th,
                    "pct_change_75th": pct_75th,
                }
            self.sensitivity[loc_key] = loc_sens

    def get_parameter_summary(self) -> str:
        """Return a rich markdown summary for LLM agent consumption."""
        lines: list[str] = []
        lines.append("# Vitamin A Supplementation (VAS) CEA — Parameter Summary\n")

        lines.append("## Model Overview\n")
        lines.append(
            "This CEA estimates the cost-effectiveness of vitamin A supplementation "
            "programs funded via Helen Keller International (HKI) and Nutrition "
            "International (NI). It covers 37 location columns: 8 HKI countries, "
            "21 Nigerian locations (20 states + FCT, HKI), and 4 NI countries. "
            "The model calculates cost-effectiveness as multiples of GiveDirectly's "
            "unconditional cash transfer program.\n"
        )

        lines.append("## Key Shared Parameters\n")
        lines.append(f"- **Grant size:** ${self.grant_size:,.0f}")
        lines.append(f"- **Supplementation rounds per year:** {self.rounds_per_year:.0f}")
        lines.append(f"- **Moral value of averting an under-5 death:** {self.moral_value_under5:.5f} UoV")
        lines.append(
            f"- **Proportion of post-neonatal deaths among 1-5 month-olds:** "
            f"{self.prop_postneonatal_1_5mo:.2f} (used to estimate VAS-eligible "
            f"6-59 month mortality from total under-5 mortality)"
        )
        lines.append("")

        lines.append("## Per-Location Parameters\n")
        lines.append(
            "| Location | Grantee | CE Multiple (x-cash) | Cost/Supplement | "
            "VAD Prevalence | VAD Survey Year | Mortality Rate (6-59mo) |"
        )
        lines.append("|---|---|---|---|---|---|---|")

        for key in list(self._LOC_COLS.keys()) + ["nigeria"]:
            loc = self.locations[key]
            if key == "nigeria":
                ce_str = (
                    f"{loc.ce_multiple_mean:.2f} "
                    f"(range: {loc.ce_multiple_min:.2f}-{loc.ce_multiple_max:.2f}, "
                    f"{loc.n_states} locations)"
                )
            else:
                ce_str = f"{loc.ce_multiple:.2f}"
            year_str = str(loc.vad_survey_year) if loc.vad_survey_year > 0 else "N/A"
            vad_str = f"{loc.vad_prevalence:.1%}" if loc.vad_prevalence > 0 else "N/A"
            mort_str = f"{loc.mortality_rate_6_59mo:.6f}" if loc.mortality_rate_6_59mo > 0 else "N/A"
            lines.append(
                f"| {loc.name} | {loc.grantee} | {ce_str} | "
                f"${loc.cost_per_supplement:.2f} | {vad_str} | {year_str} | {mort_str} |"
            )
        lines.append("")

        lines.append("### Nigerian State/FCT Breakdown\n")
        lines.append("| State/Territory | CE Multiple (x-cash) |")
        lines.append("|---|---|")
        for state_name, ce in sorted(self.nigeria_states.items(), key=lambda x: -x[1]):
            lines.append(f"| {state_name} | {ce:.2f} |")
        lines.append("")

        lines.append("## Adjustment Factors (from Simple CEA)\n")
        lines.append(
            "| Location | Additional Benefits | Grantee Factors | Leverage | Funging |"
        )
        lines.append("|---|---|---|---|---|")
        for key in list(self._LOC_COLS.keys()) + ["nigeria"]:
            loc = self.locations[key]
            lines.append(
                f"| {loc.name} | {loc.adj_additional_benefits:.3f} | "
                f"{loc.adj_grantee_factors:+.3f} | "
                f"{loc.adj_leverage:+.4f} | {loc.adj_funging:+.4f} |"
            )
        lines.append("")

        lines.append("## Sensitivity Analysis — Highest-Variance Parameters\n")
        lines.append(
            "Percentage change in final CE when each parameter moves to its "
            "25th or 75th percentile value:\n"
        )
        for loc_key, loc_name in [("burkina_faso", "Burkina Faso"), ("angola", "Angola")]:
            if loc_key not in self.sensitivity:
                continue
            sens = self.sensitivity[loc_key]
            lines.append(f"### {loc_name}\n")
            lines.append("| Parameter | 25th %ile Change | 75th %ile Change |")
            lines.append("|---|---|---|")
            for param, vals in sens.items():
                label = param.replace("_", " ").title()
                lines.append(
                    f"| {label} | {vals['pct_change_25th']:+.1%} | "
                    f"{vals['pct_change_75th']:+.1%} |"
                )
            lines.append("")

        lines.append("## Structural Concerns\n")

        stale_surveys: list[str] = []
        for key, loc in self.locations.items():
            if 0 < loc.vad_survey_year < 2015:
                stale_surveys.append(
                    f"- **{loc.name}**: VAD survey from {loc.vad_survey_year} "
                    f"({2024 - loc.vad_survey_year} years old). Estimated 2021 "
                    f"prevalence ({loc.estimated_vad_2021:.1%}) is extrapolated "
                    f"from proxy indicators, not direct measurement."
                )
        if stale_surveys:
            lines.append("### Stale VAD Prevalence Data\n")
            lines.extend(stale_surveys)
            lines.append("")

        lines.append("### DUMMYFUNCTION Cells in Inputs Sheet\n")
        lines.append(
            "Inputs rows 24-27 (mortality data) use `IFERROR(DUMMYFUNCTION(...))` "
            "formulas that resolve to numeric fallback values. These values match "
            "the raw GBD data in the Counterfactual mortality sheet. The pipeline "
            "reads from Counterfactual mortality directly.\n"
        )

        lines.append("### Nigerian State Disaggregation Asymmetry\n")
        ng = self.locations["nigeria"]
        lines.append(
            f"21 Nigerian locations (20 states + FCT) have individual columns "
            f"but share many national-level parameters. CE multiples range from "
            f"{ng.ce_multiple_min:.2f}x to {ng.ce_multiple_max:.2f}x, a "
            f"{ng.ce_multiple_max / ng.ce_multiple_min:.0f}x spread driven "
            f"primarily by cost and mortality differences. See the per-state "
            f"table above for the full breakdown.\n"
        )

        lines.append("### Static Mortality Effect\n")
        lines.append(
            "The VAS mortality effect estimate is applied as a single scalar "
            "across all locations. This is the highest-variance parameter in the "
            "sensitivity analysis (25th/75th percentile: -80%/+75% change in CE). "
            "Temporal trends in disease burden are not modeled.\n"
        )

        lines.append("### VAD Deficiency as Mediator\n")
        lines.append(
            "The external validity adjustment uses proxy indicators (stunting, "
            "wasting, poverty rates) rather than direct VAD measurement for most "
            "locations. Changes in these proxies are weighted equally (1/3 each) "
            "to estimate changes in VAD prevalence over time.\n"
        )

        return "\n".join(lines)

    def compute_cost_effectiveness(self, program_key: str, **overrides: float) -> float:
        """Return the pre-computed CE multiple for a location.

        Raises NotImplementedError if overrides are passed — VASCEA reads
        pre-computed values and cannot recompute with different parameters.
        """
        if overrides:
            raise NotImplementedError(
                "VASCEA reads pre-computed CE multiples and cannot recompute "
                "with overrides. A full formula-chain implementation would be "
                "needed for parameter sweeps."
            )
        return self.locations[program_key].ce_multiple

    def detect_cap_binding(self, program_key: str, **overrides: float) -> bool:
        """No plausibility cap concept in the VAS model."""
        return False

    def run_sensitivity(
        self,
        program_key: str,
        parameter_name: str,
        low: float,
        central: float,
        high: float,
    ) -> dict[str, Any]:
        """Return pre-computed sensitivity data from the Sensitivity analysis sheet.

        Unlike WaterCEA/ITNCEA, this returns spreadsheet-sourced percentage changes
        rather than recomputing CE at each parameter value.
        """
        if program_key in self.sensitivity and parameter_name in self.sensitivity[program_key]:
            sens = self.sensitivity[program_key][parameter_name]
            baseline = self.locations[program_key].ce_multiple
            return {
                "parameter_name": parameter_name,
                "baseline": baseline,
                "cap_binds_baseline": False,
                "low": baseline * (1 + sens["pct_change_25th"]),
                "central": baseline,
                "high": baseline * (1 + sens["pct_change_75th"]),
                "pct_change_low": sens["pct_change_25th"] * 100,
                "pct_change_central": 0.0,
                "pct_change_high": sens["pct_change_75th"] * 100,
                "cap_binds_low": False,
                "cap_binds_central": False,
                "cap_binds_high": False,
            }
        return {
            "parameter_name": parameter_name,
            "baseline": self.locations.get(program_key, self.locations["burkina_faso"]).ce_multiple,
        }
