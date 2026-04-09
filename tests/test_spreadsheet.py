"""Tests for the CEA spreadsheet reader against known WaterCEA.xlsx values."""
from __future__ import annotations

import math
from pathlib import Path

import pytest

from pipeline.spreadsheet import CEAReader, ITNCEA, MalariaCEA, VASCEA, WaterCEA

DATA_PATH = Path(__file__).parent.parent / "data" / "GW" / "WaterCEA.xlsx"
ITN_DATA_PATH = Path(__file__).parent.parent / "data" / "GW" / "InsecticideCEA.xlsx"
MALARIA_DATA_PATH = Path(__file__).parent.parent / "data" / "GW" / "MalariaCEA.xlsx"
VAS_DATA_PATH = Path(__file__).parent.parent / "data" / "VASCEA.xlsx"

_VAS_EXPECTED_SHEETS = {
    "Key", "Main CEA", "Counterfactual mortality", "External validity",
    "LeverageFunging", "GBD estimates", "Inputs", "Simple CEA",
    "Sensitivity analysis",
}


@pytest.fixture(scope="module")
def cea() -> WaterCEA:
    """Load the CEA once for all tests in the module."""
    return WaterCEA(DATA_PATH)


# -----------------------------------------------------------------------
# TestWaterCEAReading
# -----------------------------------------------------------------------

class TestWaterCEAReading:
    """Verify that parameters are correctly read from the spreadsheet."""

    def test_pooled_relative_risk(self, cea: WaterCEA) -> None:
        assert cea.relative_risk == pytest.approx(0.8638932195, rel=1e-6)

    def test_internal_validity_under5(self, cea: WaterCEA) -> None:
        assert cea.internal_validity_under5 == pytest.approx(0.7957578162, rel=1e-6)

    def test_internal_validity_over5(self, cea: WaterCEA) -> None:
        assert cea.internal_validity_over5 == pytest.approx(0.504149833, rel=1e-6)

    def test_external_validity_ilc_kenya(self, cea: WaterCEA) -> None:
        assert cea.programs["ilc_kenya"].external_validity == pytest.approx(
            1.213858014, rel=1e-6
        )

    def test_external_validity_dsw_kenya(self, cea: WaterCEA) -> None:
        assert cea.programs["dsw_kenya"].external_validity == pytest.approx(
            0.5582511733, rel=1e-6
        )

    def test_plausibility_cap_ilc_kenya(self, cea: WaterCEA) -> None:
        assert cea.programs["ilc_kenya"].plausibility_cap == pytest.approx(0.109, rel=1e-6)

    def test_plausibility_cap_dsw_kenya(self, cea: WaterCEA) -> None:
        assert cea.programs["dsw_kenya"].plausibility_cap == pytest.approx(0.056, rel=1e-6)

    def test_baseline_ce_ilc_kenya(self, cea: WaterCEA) -> None:
        ce = cea.compute_cost_effectiveness("ilc_kenya")
        assert ce == pytest.approx(7.60245168, rel=1e-3)

    def test_baseline_ce_dsw_kenya(self, cea: WaterCEA) -> None:
        ce = cea.compute_cost_effectiveness("dsw_kenya")
        assert ce == pytest.approx(4.421618553, rel=1e-3)

    def test_baseline_ce_dsw_uganda(self, cea: WaterCEA) -> None:
        ce = cea.compute_cost_effectiveness("dsw_uganda")
        assert ce == pytest.approx(7.015706986, rel=1e-3)

    def test_baseline_ce_dsw_malawi(self, cea: WaterCEA) -> None:
        ce = cea.compute_cost_effectiveness("dsw_malawi")
        assert ce == pytest.approx(8.657297146, rel=1e-3)

    def test_cap_binds_ilc_kenya(self, cea: WaterCEA) -> None:
        assert cea.detect_cap_binding("ilc_kenya") is True

    def test_list_programs(self, cea: WaterCEA) -> None:
        expected = {"ilc_kenya", "dsw_kenya", "dsw_uganda", "dsw_malawi"}
        assert set(cea.programs.keys()) == expected

    def test_parameter_summary(self, cea: WaterCEA) -> None:
        summary = cea.get_parameter_summary()
        assert len(summary) > 0
        assert "relative" in summary.lower() or "Relative" in summary


# -----------------------------------------------------------------------
# TestSensitivityAnalysis
# -----------------------------------------------------------------------

class TestSensitivityAnalysis:
    """Verify sensitivity analysis behaviour, especially cap binding."""

    def test_weaker_rr_lowers_ce(self, cea: WaterCEA) -> None:
        """With RR=0.90 (weaker effect), ILC Kenya CE should drop meaningfully.

        At RR=0.90 the plausibility cap no longer binds, but the under-5
        initial estimate is still below the cap so the overall effect is
        moderate (~5%). We verify CE drops by at least 3%.
        """
        baseline = cea.compute_cost_effectiveness("ilc_kenya")
        weaker = cea.compute_cost_effectiveness("ilc_kenya", relative_risk=0.90)
        pct_drop = ((baseline - weaker) / baseline) * 100
        assert pct_drop > 3, f"Expected >3% drop, got {pct_drop:.1f}%"

    def test_cap_binding_detection(self, cea: WaterCEA) -> None:
        """ILC Kenya cap should bind at baseline."""
        assert cea.detect_cap_binding("ilc_kenya") is True

    def test_cap_not_binding_with_weaker_rr(self, cea: WaterCEA) -> None:
        """ILC Kenya cap should NOT bind at RR=0.90."""
        assert cea.detect_cap_binding("ilc_kenya", relative_risk=0.90) is False

    def test_run_sensitivity_returns_expected_keys(self, cea: WaterCEA) -> None:
        result = cea.run_sensitivity(
            "ilc_kenya", "relative_risk", low=0.90, central=0.8639, high=0.80
        )
        expected_keys = {
            "parameter_name",
            "baseline",
            "low",
            "central",
            "high",
            "pct_change_low",
            "pct_change_central",
            "pct_change_high",
            "cap_binds_baseline",
            "cap_binds_low",
            "cap_binds_central",
            "cap_binds_high",
        }
        assert expected_keys.issubset(set(result.keys()))

    def test_stronger_rr_capped(self, cea: WaterCEA) -> None:
        """Strengthening RR when cap binds should NOT change under-5 mortality
        contribution, so overall CE change should be very small for ILC Kenya."""
        baseline = cea.compute_cost_effectiveness("ilc_kenya")
        # Stronger RR (lower value = bigger effect)
        stronger = cea.compute_cost_effectiveness("ilc_kenya", relative_risk=0.80)
        # When cap binds, under-5 mortality is unchanged. Over-5 mortality
        # changes through cap_ratio, and development effects stay the same
        # (IV and EV unchanged). Medical costs stay the same.
        # So the change should be very small (only through over-5 cap_ratio).
        pct_change = abs((stronger - baseline) / baseline) * 100
        # The cap ratio changes: with stronger RR, initial_estimate increases,
        # so cap_ratio = cap/initial decreases, which actually LOWERS over-5 UV.
        # Net effect: CE should barely change or decrease slightly.
        assert pct_change < 5, f"Expected <5% change when cap binds, got {pct_change:.1f}%"


# -----------------------------------------------------------------------
# ITN CEA fixtures
# -----------------------------------------------------------------------


@pytest.fixture(scope="module")
def itn_cea() -> ITNCEA:
    """Load the ITN CEA once for all tests in the module."""
    return ITNCEA(ITN_DATA_PATH)


# -----------------------------------------------------------------------
# TestITNCEAReading
# -----------------------------------------------------------------------


class TestITNCEAReading:
    """Verify that ITNCEA reads parameters and computes baseline CE correctly."""

    # Validation targets from Simple CEA R38, cols I-P
    EXPECTED_CE = {
        "chad": 4.82,
        "drc": 14.63,
        "guinea": 22.79,
        "nigeria_gf": 16.78,
        "nigeria_pmi": 13.25,
        "south_sudan": 7.16,
        "togo": 8.81,
        "uganda": 15.60,
    }

    @pytest.mark.parametrize("program,expected", list(EXPECTED_CE.items()))
    def test_baseline_ce(self, itn_cea: ITNCEA, program: str, expected: float) -> None:
        ce = itn_cea.compute_cost_effectiveness(program)
        assert ce == pytest.approx(expected, rel=1e-2), (
            f"{program}: expected {expected}, got {ce:.4f}"
        )

    def test_list_programs(self, itn_cea: ITNCEA) -> None:
        expected = {
            "chad", "drc", "guinea", "nigeria_gf",
            "nigeria_pmi", "south_sudan", "togo", "uganda",
        }
        assert set(itn_cea.programs.keys()) == expected

    def test_shared_params(self, itn_cea: ITNCEA) -> None:
        assert itn_cea.incidence_reduction == pytest.approx(0.45, rel=1e-6)
        assert itn_cea.internal_validity == pytest.approx(-0.05, rel=1e-6)
        assert itn_cea.external_validity == pytest.approx(-0.05, rel=1e-6)
        assert itn_cea.indirect_deaths_multiplier == pytest.approx(0.75, rel=1e-6)
        assert itn_cea.moral_weight_u5 == pytest.approx(116.25262, rel=1e-4)

    def test_insecticide_resistance_varies(self, itn_cea: ITNCEA) -> None:
        """Resistance varies dramatically across countries."""
        resistances = [itn_cea.programs[k].insecticide_resistance for k in itn_cea.PROGRAMS]
        assert min(resistances) < -0.5  # South Sudan or Chad
        assert max(resistances) > -0.12  # DRC

    def test_parameter_summary(self, itn_cea: ITNCEA) -> None:
        summary = itn_cea.get_parameter_summary()
        assert len(summary) > 0
        assert "insecticide resistance" in summary.lower()
        assert "Chad" in summary

    def test_cap_never_binds(self, itn_cea: ITNCEA) -> None:
        for key in itn_cea.PROGRAMS:
            assert itn_cea.detect_cap_binding(key) is False


# -----------------------------------------------------------------------
# TestITNCEASensitivity
# -----------------------------------------------------------------------


class TestITNCEASensitivity:
    """Verify sensitivity analysis for ITN CEA."""

    def test_weaker_incidence_reduction_lowers_ce(self, itn_cea: ITNCEA) -> None:
        baseline = itn_cea.compute_cost_effectiveness("drc")
        weaker = itn_cea.compute_cost_effectiveness("drc", incidence_reduction=0.30)
        assert weaker < baseline, "Weaker incidence reduction should lower CE"

    def test_worse_resistance_lowers_ce(self, itn_cea: ITNCEA) -> None:
        baseline = itn_cea.compute_cost_effectiveness("drc")
        worse = itn_cea.compute_cost_effectiveness("drc", insecticide_resistance=-0.50)
        assert worse < baseline, "Worse insecticide resistance should lower CE"

    def test_higher_indirect_multiplier_raises_ce(self, itn_cea: ITNCEA) -> None:
        baseline = itn_cea.compute_cost_effectiveness("chad")
        higher = itn_cea.compute_cost_effectiveness("chad", indirect_deaths_multiplier=1.0)
        assert higher > baseline, "Higher indirect multiplier should raise CE"

    def test_run_sensitivity_returns_expected_keys(self, itn_cea: ITNCEA) -> None:
        result = itn_cea.run_sensitivity(
            "chad", "incidence_reduction", low=0.30, central=0.45, high=0.60
        )
        expected_keys = {
            "parameter_name", "baseline",
            "low", "central", "high",
            "pct_change_low", "pct_change_central", "pct_change_high",
            "cap_binds_baseline", "cap_binds_low", "cap_binds_central", "cap_binds_high",
        }
        assert expected_keys.issubset(set(result.keys()))

    def test_run_sensitivity_nonempty(self, itn_cea: ITNCEA) -> None:
        result = itn_cea.run_sensitivity(
            "guinea", "insecticide_resistance", low=-0.40, central=-0.207, high=-0.10
        )
        assert result["baseline"] > 0
        assert result["pct_change_low"] != 0.0

    def test_sensitivity_direction_incidence(self, itn_cea: ITNCEA) -> None:
        """Lowering incidence_reduction should lower CE (negative pct_change_low)."""
        result = itn_cea.run_sensitivity(
            "uganda", "incidence_reduction", low=0.30, central=0.45, high=0.60
        )
        assert result["pct_change_low"] < 0, "Lower incidence reduction → lower CE"
        assert result["pct_change_high"] > 0, "Higher incidence reduction → higher CE"


# -----------------------------------------------------------------------
# VAS CEA fixtures and tests
# -----------------------------------------------------------------------


@pytest.fixture(scope="module")
def vas_cea() -> VASCEA:
    """Load the VAS CEA once for all tests in the module."""
    return VASCEA(VAS_DATA_PATH)


class TestVASCEAReading:
    """Verify that VAS parameters are correctly read from the spreadsheet."""

    def test_required_sheets_present(self) -> None:
        """All 5 sheets that VASCEA reads must exist in the workbook."""
        import openpyxl
        wb = openpyxl.load_workbook(str(VAS_DATA_PATH), data_only=True)
        actual = set(wb.sheetnames)
        wb.close()
        missing = _VAS_EXPECTED_SHEETS - actual
        assert not missing, f"Missing sheets: {missing}"

    def test_locations_loaded(self, vas_cea: VASCEA) -> None:
        """12 non-Nigerian locations + 1 aggregated Nigeria = 13 entries."""
        assert "burkina_faso" in vas_cea.locations
        assert "angola" in vas_cea.locations
        assert "nigeria" in vas_cea.locations
        assert len(vas_cea.locations) == 13

    def test_ce_multiple_burkina_faso(self, vas_cea: VASCEA) -> None:
        """Cross-check: Burkina Faso CE from Simple CEA row 37 col I."""
        loc = vas_cea.locations["burkina_faso"]
        assert loc.ce_multiple == pytest.approx(6.851640881, rel=1e-4)

    def test_ce_multiple_niger(self, vas_cea: VASCEA) -> None:
        """Niger has the highest non-Nigerian CE multiple at ~79x."""
        loc = vas_cea.locations["niger"]
        assert loc.ce_multiple == pytest.approx(79.13069972, rel=1e-4)

    def test_ce_multiple_angola(self, vas_cea: VASCEA) -> None:
        """Angola (Nutrition International) at ~3.69x."""
        loc = vas_cea.locations["angola"]
        assert loc.ce_multiple == pytest.approx(3.690803672, rel=1e-4)

    def test_ce_matches_main_cea_header(self, vas_cea: VASCEA) -> None:
        """Hand-verify: Burkina Faso CE matches Main CEA row 4 (the header echo)."""
        import openpyxl
        wb = openpyxl.load_workbook(str(VAS_DATA_PATH), data_only=True)
        main_cea_val = float(wb["Main CEA"].cell(row=4, column=9).value)
        wb.close()
        assert vas_cea.locations["burkina_faso"].ce_multiple == pytest.approx(
            main_cea_val, rel=1e-6
        )

    def test_cost_per_supplement_burkina_faso(self, vas_cea: VASCEA) -> None:
        loc = vas_cea.locations["burkina_faso"]
        assert loc.cost_per_supplement == pytest.approx(1.5410928, rel=1e-4)

    def test_vad_survey_year_angola(self, vas_cea: VASCEA) -> None:
        """Angola's VAD survey is from 1999 — must be flagged as stale."""
        loc = vas_cea.locations["angola"]
        assert loc.vad_survey_year == 1999

    def test_vad_survey_year_chad(self, vas_cea: VASCEA) -> None:
        loc = vas_cea.locations["chad"]
        assert loc.vad_survey_year == 2008

    def test_vad_survey_year_togo_missing(self, vas_cea: VASCEA) -> None:
        """Togo has no VAD survey (cell is '-')."""
        loc = vas_cea.locations["togo"]
        assert loc.vad_survey_year == 0

    def test_nigeria_is_aggregated(self, vas_cea: VASCEA) -> None:
        """Nigeria aggregates 21 locations (20 states + FCT) into min/max/mean."""
        loc = vas_cea.locations["nigeria"]
        assert loc.n_states == 21
        assert loc.ce_multiple_min == pytest.approx(1.3704, rel=1e-2)
        assert loc.ce_multiple_max == pytest.approx(38.2464, rel=1e-2)
        assert loc.ce_multiple_mean == pytest.approx(6.9547, rel=1e-2)

    def test_nigeria_initial_ce_differs_from_final(self, vas_cea: VASCEA) -> None:
        """Initial CE (before adjustments) must be computed separately, not copied from final."""
        loc = vas_cea.locations["nigeria"]
        # Mean initial CE across 21 states is 5.4081, vs final mean 6.9547
        assert loc.initial_ce_multiple == pytest.approx(5.4081, rel=1e-2)
        assert loc.initial_ce_multiple != pytest.approx(loc.ce_multiple, rel=1e-2)

    def test_nigeria_states_exposed(self, vas_cea: VASCEA) -> None:
        """Individual state CE values should be accessible for the summary."""
        assert len(vas_cea.nigeria_states) == 21
        assert "Sokoto" in vas_cea.nigeria_states
        assert vas_cea.nigeria_states["Sokoto"] == pytest.approx(38.2464, rel=1e-2)
        assert "FCT (Abuja)" in vas_cea.nigeria_states

    def test_shared_params(self, vas_cea: VASCEA) -> None:
        assert vas_cea.grant_size == pytest.approx(1_000_000, rel=1e-6)
        assert vas_cea.rounds_per_year == pytest.approx(2.0, rel=1e-6)
        assert vas_cea.moral_value_under5 == pytest.approx(118.73259, rel=1e-4)

    def test_sensitivity_mortality_effect(self, vas_cea: VASCEA) -> None:
        """Mortality effect is the highest-variance parameter: -80%/+75%."""
        sens = vas_cea.sensitivity["burkina_faso"]
        mort_effect = sens["effect_of_vas_on_mortality"]
        assert mort_effect["pct_change_25th"] == pytest.approx(-0.80, rel=1e-2)
        assert mort_effect["pct_change_75th"] == pytest.approx(0.75, rel=1e-2)

    def test_compute_ce_rejects_overrides(self, vas_cea: VASCEA) -> None:
        """compute_cost_effectiveness must refuse overrides (no formula chain)."""
        with pytest.raises(NotImplementedError):
            vas_cea.compute_cost_effectiveness("burkina_faso", relative_risk=0.5)

    def test_compute_ce_without_overrides(self, vas_cea: VASCEA) -> None:
        """Without overrides, returns the pre-computed CE multiple."""
        ce = vas_cea.compute_cost_effectiveness("burkina_faso")
        assert ce == pytest.approx(6.851640881, rel=1e-4)

    def test_parameter_summary_completeness(self, vas_cea: VASCEA) -> None:
        """Summary must contain all 13 location names and key structural flags."""
        summary = vas_cea.get_parameter_summary()
        assert len(summary) > 500

        # All 13 locations present
        for name in [
            "Burkina Faso", "Cameroon", "Cote d'Ivoire", "DRC", "Guinea",
            "Madagascar", "Mali", "Niger", "Angola", "Chad", "Togo", "Uganda",
            "Nigeria",
        ]:
            assert name in summary, f"Missing location: {name}"

        # Stale survey flags
        assert "1999" in summary  # Angola
        assert "2008" in summary  # Chad

        # Nigeria range string
        import re
        assert re.search(r"range: \d+\.\d+.+\d+\.\d+", summary), \
            "Nigeria range string missing from summary"

        # Per-state breakdown present
        assert "Sokoto" in summary
        assert "Anambra" in summary


# -----------------------------------------------------------------------
# TestCEAReaderInterface
# -----------------------------------------------------------------------

@pytest.fixture(scope="module")
def all_ceas() -> dict[str, CEAReader]:
    """Instantiate every concrete CEA class once for interface testing."""
    return {
        "water": WaterCEA(DATA_PATH),
        "itns": ITNCEA(ITN_DATA_PATH),
        "smc": MalariaCEA(MALARIA_DATA_PATH),
        "vas": VASCEA(VAS_DATA_PATH),
    }


class TestCEAReaderInterface:
    """Verify every CEA class conforms to the CEAReader Protocol.

    This catches missing-attribute bugs at test time. The original motivation
    was a mid-run crash on VAS at quantifier critique 5/31 caused by VASCEA
    missing the PROGRAMS class attribute that parse_quantifier_output iterates.
    """

    @pytest.mark.parametrize("name", ["water", "itns", "smc", "vas"])
    def test_isinstance_check(self, all_ceas: dict[str, CEAReader], name: str) -> None:
        """Each CEA class is recognized as a CEAReader by runtime isinstance."""
        cea = all_ceas[name]
        assert isinstance(cea, CEAReader), (
            f"{type(cea).__name__} does not satisfy the CEAReader Protocol. "
            f"Check that all required members are present."
        )

    @pytest.mark.parametrize("name", ["water", "itns", "smc", "vas"])
    def test_programs_is_tuple_of_strings(
        self, all_ceas: dict[str, CEAReader], name: str
    ) -> None:
        cea = all_ceas[name]
        assert isinstance(cea.PROGRAMS, tuple), (
            f"{type(cea).__name__}.PROGRAMS must be a tuple"
        )
        assert len(cea.PROGRAMS) > 0, f"{type(cea).__name__}.PROGRAMS is empty"
        assert all(isinstance(p, str) for p in cea.PROGRAMS), (
            f"{type(cea).__name__}.PROGRAMS must contain only strings"
        )

    @pytest.mark.parametrize("name", ["water", "itns", "smc", "vas"])
    def test_get_parameter_summary_returns_nonempty_string(
        self, all_ceas: dict[str, CEAReader], name: str
    ) -> None:
        cea = all_ceas[name]
        summary = cea.get_parameter_summary()
        assert isinstance(summary, str)
        assert len(summary) > 500, (
            f"{type(cea).__name__}.get_parameter_summary() returned only "
            f"{len(summary)} chars; expected > 500"
        )

    @pytest.mark.parametrize("name", ["water", "itns", "smc", "vas"])
    def test_compute_cost_effectiveness_baseline(
        self, all_ceas: dict[str, CEAReader], name: str
    ) -> None:
        """compute_cost_effectiveness returns a float for the first program key."""
        cea = all_ceas[name]
        first_program = cea.PROGRAMS[0]
        result = cea.compute_cost_effectiveness(first_program)
        assert isinstance(result, (int, float))
        assert result > 0, (
            f"{type(cea).__name__}.compute_cost_effectiveness({first_program!r}) "
            f"returned {result}; expected > 0"
        )

    @pytest.mark.parametrize("name", ["water", "itns", "smc", "vas"])
    def test_detect_cap_binding_returns_bool(
        self, all_ceas: dict[str, CEAReader], name: str
    ) -> None:
        cea = all_ceas[name]
        first_program = cea.PROGRAMS[0]
        result = cea.detect_cap_binding(first_program)
        assert isinstance(result, bool)

    def test_vas_compute_with_overrides_raises(
        self, all_ceas: dict[str, CEAReader]
    ) -> None:
        """VASCEA explicitly fails loud on overrides since it can't recompute."""
        vas = all_ceas["vas"]
        first_program = vas.PROGRAMS[0]
        with pytest.raises(NotImplementedError):
            vas.compute_cost_effectiveness(first_program, some_param=1.0)
